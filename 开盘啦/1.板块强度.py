from email import header
from time import sleep
from openpyxl import Workbook,load_workbook

from openpyxl.styles import Alignment
import datetime
import requests


# print(json)
# 格式化数据
def format_number(num):
    yi = 10 ** 8
    w = 10 ** 4
    q = 10 ** 3
    if(num >= yi or -num >=yi):
        return f'{round(num/yi, 2)}亿'
    elif(num >= w or -num >=w):
        return f'{round(num/w, 2)}万'
    else:
        return f'{round(num/q, 2)}千'
# 保存在excle
def save_to_excle(data_info):
    date = data_info['date']
    # file_name = f'{date}.xlsx'
    file_name = 'demo.xlsx'
    header_arr = ['板块名称','强度']
    
   
    # 已经创建这个文件时
    wb = load_workbook(file_name)
    wb.save(file_name)
    ws = wb.active
    
    # 最大行
    max_row = ws.max_row
    # 最大列
    max_col = ws.max_column
    print(max_row,max_col)

    # 如果是已经写入
    if(max_col > 1):
        
        offset_col =  max_col + 2
        # 向右写入
        ws.cell(1,offset_col,data_info['date'])

        for col in range(offset_col,(offset_col+len(data_info['list'][0]))):
            ws.cell(2,col,header_arr[col-(offset_col)])

        for row in range(3,max_row+1):
            for col in range(offset_col,(offset_col+len(data_info['list'][0]))):
                ws.cell(row = row,column = col,value = data_info['list'][row-3][col-(offset_col)])
        wb.save(file_name)   
        print("保存成功")
    else:
        # 如果第一次写入
        ws.append([data_info['date']])
        ws.append(header_arr)
        for row in data_info['list']:
            ws.append(row)
        wb.save(file_name)
        print("保存成功")

# 根据日期获取数据并保存
def getdate_form_day(day):
    header = {
        'Host':'apphis.longhuvip.com',
        'Content-Type':'application/x-www-form-urlencoded; charset=utf-8',
        'Connection':'keep-alive',
        'Accept':'*/*',
        'User-Agent':'Mozilla/5.0 (Linux; Android 7.1.2; SM-N976N Build/QP1A.190711.020; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/92.0.4515.131 Mobile Safari/537.36;kaipanla 5.6.0.2',
        'Accept-Language':'zh-Hans-CN;q=1.0, en-CN;q=0.9',
        'Content-Length':'125',
        'Accept-Encoding':'gzip;q=1.0, compress;q=0.5'
    }
    data = {
        'Date':day, #2022-07-25
        "Index":"0",
        'Order':'1',
        'PhoneOSNew':'2',
        "Type":'1',
        "VerSion":'5.6.0.3',
        "ZSType":"7",
        "a":"RealRankingInfo",
        'apiv':"w30",
        "c":"ZhiShuRanking",
        "st":'20'
    }
    res = requests.post("https://apphis.longhuvip.com/w1/api/index.php",headers=header,data=data)
    json = res.json()

    data['Order'] = 0
    res = requests.post("https://apphis.longhuvip.com/w1/api/index.php",headers=header,data=data)
    json2 = res.json()
    
    for i in json2['list']:
        json['list'].append(i)

    info = {'date':json['Day'][0],'list':[]}
    for i in json['list']:
        arr_item = [i[1], i[2]]
        info['list'].append(arr_item)
    print(len(info['list']))
    save_to_excle(info)

# 获取当天数据并保存
def getdate_today():
    header = {
        'Host':'apphq.longhuvip.com',
        'Content-Type':'application/x-www-form-urlencoded; charset=utf-8',
        'Connection':'keep-alive',
        'Accept':'*/*',
        'User-Agent':'Mozilla/5.0 (Linux; Android 7.1.2; SM-N976N Build/QP1A.190711.020; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/92.0.4515.131 Mobile Safari/537.36;kaipanla 5.6.0.2',
        'Accept-Language':'zh-Hans-CN;q=1.0, en-CN;q=0.9',
        'Content-Length':'109',
        'Accept-Encoding':'gzip;q=1.0, compress;q=0.5'
    }
    data = {
        "Index":"0",
        'Order':'1',
        'PhoneOSNew':'2',
        "Type":'1',
        "VerSion":'5.6.0.3',
        "ZSType":"7",
        "a":"RealRankingInfo",
        'apiv':"w30",
        "c":"ZhiShuRanking",
        "st":'20'
    }
    res = requests.post("https://apphq.longhuvip.com/w1/api/index.php",headers=header,data=data)
    json = res.json()
    data['Order'] = 0
    res = requests.post("https://apphq.longhuvip.com/w1/api/index.php",headers=header,data=data)
    json2 = res.json()
    
    for i in json2['list']:
        json['list'].append(i)

    info = {'date':json['Day'][0],'list':[]}
    for i in json['list']:
        arr_item = [i[1], i[2]]
        info['list'].append(arr_item)
    print(len(info['list']))
    save_to_excle(info)



# now = datetime.datetime.now()
# today = f'{now.year}-{now.month}-{now.day}'
# getdate_form_day(today)  

getdate_today()




